from openpyxl import load_workbook
wb = load_workbook('C:\\Direccion_del_archivo\\Nombre_archivo.xlsx') #Agregar la direccion del archivo y el nombre del archivo del usuario
ws = wb.active
columna = 1
fila = 2 
while ws.cell(row=fila, column=1).value is not None:
    fila += 1
ultima_fila = fila
def ingresar_datos(fila, columna):
    datos_correctos = False 
    while not datos_correctos:
        fecha = input("Ingrese la fecha del movimiento (dd/mm/aaaa): ")
        tipo_movimiento = input("Ingrese tipo de movimiento (ingreso/gasto/ahorro): ")
        categoria = input("Ingrese la categoria del movimiento(si es ahorro, ingrese 'ahorro'): ")
        especificacion = input("Ingrese la especificacion del movimiento(si es ahorro, ingrese 'ahorro'): ")
        monto = input("Ingrese el monto del movimiento: ")
        print("¿Los datos ingresados son correctos?:")
        print(f"Fecha: {fecha}")
        print(f"Tipo de movimiento: {tipo_movimiento}")
        print(f"Categoría: {categoria}")
        print(f"Especificación: {especificacion}")
        print(f"Monto: {monto}")
        confirmacion_datos = input("¿Los datos son correctos?(si/no): ").lower()
        if confirmacion_datos == 'si':
            datos_correctos = True
    ws.cell(row=fila, column=columna).value = fecha
    ws.cell(row=fila, column=columna+1).value = tipo_movimiento
    ws.cell(row=fila, column=columna+2).value = categoria
    ws.cell(row=fila, column=columna+3).value = especificacion
    ws.cell(row=fila, column=columna+4).value = monto
    print(f"Guardando en fila {fila}")
    wb.save('C:\\Users\\PC\\Franco\\Trabajos papá\\Registro diario - copia.xlsx')
    print("Guardado exitoso")
def consultar(listo):
    global ultima_fila
    print("¿Quiere ingresar nuevos datos?")
    respuesta = input("Ingrese 'si' para continuar o 'no' para salir: ").lower()
    if respuesta == 'si':
        ingresar_datos(ultima_fila, columna)
        ultima_fila = ultima_fila + 1
    else:
        print("Gracias por usar el programa. ¡Hasta luego!")
        return True
while True:
    listo = consultar(False)
    if listo:
        break
    